#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
decision_kit_en.py

Enhanced decision kit for Update Description 3 (EN variant).

- Works from QA files produced by build_desc3_enh.py (with desc_src_www).
- Detects:
    * GLUE_CANDIDATE: adjacent tokens where one side is short (<=2 chars)
      and concatenation looks valid; proposes glue pair → concat.
    * UNKNOWN_ABBREV: short tokens (<=4 letters) not in known sets.
- Computes row_hits (distinct rowpointer count), occurrences (raw), and where (desc3/www/both).
- Creates evidence excerpts with highlight markers:
    * glue bigram "c lip" shown as [[c][lip]]
    * abbrev "zn" shown as [[zn]]
- Outputs decisions CSV + large-font XLSX with dropdowns.
- summary/validate/apply mirror the classic decision_kit, writing to:
    configs/lines/<line>.json5, configs/global/glue.json5, configs/global/synonyms.json5
"""

import argparse, os, re, sys, glob, json, datetime
from typing import Optional, List
import pandas as pd

# --------- repo config locations ----------
QA_DIR_DEFAULT     = "rep/qa"
DECISIONS_DIR      = "decisions"
SNAP_DIR           = os.path.join(DECISIONS_DIR, "snapshots")
GLOBAL_SYNONYMS    = "configs/global/synonyms.json5"
GLOBAL_GLUE        = "configs/global/glue.json5"
LINE_FILE_TMPL     = "configs/lines/{line}.json5"
CSV_EOL            = "\r\n"

# Columns to surface prominently
BASE_COLS = [
    "issue_type",
    "pair", "concat", "abbrev",
    "row_hits", "occurrences", "where",
    "evidence_desc3", "evidence_www",
    "sample_prod", "sample_source-desc-name", "sample_attributegrp",
    "sample_descrip1", "sample_descrip2", "sample_descrip3",
    "sample_lookup",
    "count", "option1", "option2", "option3", "option4"
]
DECISION_COLS = ["decision", "approved_value", "scope_hint", "notes"]
DECISION_ALLOWED = ["accept_line", "accept_global", "ignore", "defer"]
SCOPE_ALLOWED    = ["line", "global"]

# ---------- JSON helpers ----------
def _ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)

def _now_ts() -> str:
    return datetime.datetime.now().strftime("%Y%m%d_%H%M")

def _load_json(path: str) -> dict:
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def _save_json(path: str, data: dict) -> None:
    _ensure_dir(os.path.dirname(path))
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

# ---------- file ops ----------
def _latest_en_qa(line: str, qa_dir: str = QA_DIR_DEFAULT) -> Optional[str]:
    patt = os.path.join(qa_dir, f"QA_{line}_en_*.csv")
    files = sorted(glob.glob(patt))
    return files[-1] if files else None

def _read_any(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return pd.read_excel(path, dtype=str, engine="openpyxl").fillna("")
    return pd.read_csv(path, dtype=str).fillna("")

def _write_csv(df: pd.DataFrame, path: str) -> None:
    _ensure_dir(os.path.dirname(path))
    df.to_csv(path, index=False, lineterminator=CSV_EOL)

def _ensure_decision_cols(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in BASE_COLS:
        if c not in out.columns:
            out[c] = ""
    for c in DECISION_COLS:
        if c not in out.columns:
            out[c] = "" if c != "scope_hint" else "line"
    order = [c for c in BASE_COLS if c in out.columns] + DECISION_COLS
    leftover = [c for c in out.columns if c not in order]
    return out[order + leftover]

def _xlsx_format_template(xlsx_path: str) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.title = "Decisions"

    headers = {cell.value: cell.column for cell in next(ws.iter_rows(min_row=1, max_row=1))}
    def L(name):
        idx = headers.get(name)
        return get_column_letter(idx) if idx else None

    # styling
    header_fill = PatternFill("solid", fgColor="000000")
    header_font = Font(bold=True, color="FFFFFF", size=16)
    body_font   = Font(size=14)
    center_al   = Alignment(vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_al

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = center_al

    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    for r in range(2, ws.max_row + 1):
        if r % 2 == 0:
            for cell in ws[r]:
                cell.fill = alt_fill

    widths = {
        "issue_type": 16, "pair": 28, "concat": 22, "abbrev": 18,
        "row_hits": 12, "occurrences": 14, "where": 10,
        "evidence_desc3": 48, "evidence_www": 48,
        "sample_prod": 22, "sample_source-desc-name": 26, "sample_attributegrp": 18,
        "sample_descrip1": 26, "sample_descrip2": 26, "sample_descrip3": 26,
        "sample_lookup": 20,
        "count": 10,
        "option1": 40, "option2": 32, "option3": 36, "option4": 28,
        "decision": 18, "approved_value": 24, "scope_hint": 12, "notes": 32,
    }
    for name, w in widths.items():
        c = headers.get(name)
        if c:
            ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # dropdowns
    dec_col = L("decision")
    if dec_col:
        dv1 = DataValidation(type="list", formula1='"' + ",".join(DECISION_ALLOWED) + '"', allow_blank=True)
        ws.add_data_validation(dv1)
        dv1.add(f"{dec_col}2:{dec_col}{ws.max_row}")
    scp_col = L("scope_hint")
    if scp_col:
        dv2 = DataValidation(type="list", formula1='"' + ",".join(SCOPE_ALLOWED) + '"', allow_blank=True)
        ws.add_data_validation(dv2)
        dv2.add(f"{scp_col}2:{scp_col}{ws.max_row}")

    wb.save(xlsx_path)

# ---------- known sets ----------
def _load_existing_configs(line: str):
    line_path = LINE_FILE_TMPL.format(line=line)
    line_cfg  = _load_json(line_path) or {}
    line_syn  = {**(line_cfg.get("synonyms") or {}), **(line_cfg.get("pricing_synonyms") or {})}
    line_glu  = (line_cfg.get("glue_bigrams") or {})
    glob_syn  = _load_json(GLOBAL_SYNONYMS) or {}
    glob_glu  = _load_json(GLOBAL_GLUE) or {}
    colors    = line_cfg.get("colors") or {}
    materials = line_cfg.get("materials") or {}
    uom_terms = set(line_cfg.get("uom_terms") or [])
    known = set(colors.keys()) | set(colors.values()) \
          | set(materials.keys()) | set(materials.values()) \
          | uom_terms \
          | set(line_syn.keys()) | set(line_syn.values()) \
          | set(glob_syn.keys()) | set(glob_syn.values())
    glue_map = {**glob_glu, **line_glu}
    return line_cfg, known, glue_map, glob_syn, glob_glu

# ---------- token utils ----------
def _tokens(s: str) -> List[str]:
    s = (s or "").strip()
    return [t for t in s.split() if t]

def _glue_cands(tokens: List[str], glued_pairs_set: set) -> List[tuple]:
    """
    Returns list of (pair, concat) candidates excluding already-known glues.
    Heuristic: if either side length <= 2 and concat len in [3..18].
    """
    out = []
    for i in range(len(tokens)-1):
        a, b = tokens[i], tokens[i+1]
        pair = f"{a.lower()} {b.lower()}"
        if pair in glued_pairs_set:
            continue
        if (a.isalpha() and b.isalpha()) and (len(a) <= 2 or len(b) <= 2):
            cat = (a + b).lower()
            if 3 <= len(cat) <= 18:
                out.append((pair, cat))
    return out

def _highlight_glue(text: str, pair: str) -> str:
    """Replace first two-token occurrence with [[a][b]] markers."""
    if not text:
        return ""
    a, b = pair.split(" ", 1)
    tokens = text.split()
    for i in range(len(tokens)-1):
        if tokens[i].lower() == a and tokens[i+1].lower() == b:
            tokens[i] = f"[[{tokens[i]}]]".replace("]]", "][")  # temporary to build [[a][
            tokens[i+1] = f"{tokens[i+1]}]]"
            # fix temporary marker
            tokens[i] = tokens[i].replace("][", "][")
            return " ".join(tokens)
    return text

def _highlight_abbrev(text: str, abbr: str) -> str:
    if not text:
        return ""
    tokens = text.split()
    for i in range(len(tokens)):
        if tokens[i].lower() == abbr:
            tokens[i] = f"[[{tokens[i]}]]"
            return " ".join(tokens)
    return text

# ---------- build recommendations from QA_en ----------
def cmd_derive(args):
    qa_path = args.input
    if not qa_path:
        qa_path = _latest_en_qa(args.line, qa_dir=QA_DIR_DEFAULT)
        if not qa_path:
            sys.exit(f"No enhanced QA found at {QA_DIR_DEFAULT}/QA_{args.line}_en_*.csv")

    df = _read_any(qa_path)

    # Column names in enhanced QA
    desc3_col = "descrip3_after" if "descrip3_after" in df.columns else "descrip3"
    www_col   = "desc_src_www" if "desc_src_www" in df.columns else None

    # Source columns (for samples)
    prod_col  = "prod" if "prod" in df.columns else None
    sdn_col   = "source-desc-name" if "source-desc-name" in df.columns else None
    attr_col  = "attributegrp" if "attributegrp" in df.columns else None
    d1_col    = "descrip1" if "descrip1" in df.columns else None
    d2_col    = "descrip2" if "descrip2" in df.columns else None
    d3src_col = "descrip3" if "descrip3" in df.columns else None
    look_col  = "lookupnm" if "lookupnm" in df.columns else None
    rowp_col  = "rowpointer" if "rowpointer" in df.columns else None

    line_cfg, known, glue_map, glob_syn, glob_glu = _load_existing_configs(args.line)
    glued_pairs_set = set(glue_map.keys())

    glue_rows = []  # rowpointer, where, pair, concat, evidence_desc3, evidence_www, samples...
    abbr_occ  = []  # rowpointer, where, abbrev, evidence_desc3, evidence_www, samples...
    abbr_rowu = []  # rowpointer, abbrev (per-row unique)

    for _, r in df.iterrows():
        rowp   = str(r.get(rowp_col, f"IDX:{_+1}"))
        t_desc = str(r.get(desc3_col, "") or "")
        t_www  = str(r.get(www_col, "") or "") if www_col else ""

        toks_desc = _tokens(t_desc)
        toks_www  = _tokens(t_www)

        # --- GLUE from desc3 ---
        for pair, cat in _glue_cands(toks_desc, glued_pairs_set):
            glue_rows.append((
                rowp, "desc3", pair, cat,
                _highlight_glue(t_desc, pair), "",   # evidence desc3, www
                r.get(prod_col,""), r.get(sdn_col,""), r.get(attr_col,""),
                r.get(d1_col,""), r.get(d2_col,""), r.get(d3src_col,""),
                r.get(look_col,"")
            ))
        # --- GLUE from www ---
        for pair, cat in _glue_cands(toks_www, glued_pairs_set):
            glue_rows.append((
                rowp, "www", pair, cat,
                "", _highlight_glue(t_www, pair),
                r.get(prod_col,""), r.get(sdn_col,""), r.get(attr_col,""),
                r.get(d1_col,""), r.get(d2_col,""), r.get(d3src_col,""),
                r.get(look_col,"")
            ))

        # --- ABBREV raw occurrences & per-row uniques (desc3 + www) ---
        def collect_abbrevs(tokens: List[str], where_label: str):
            for t in tokens:
                tl = t.lower()
                if tl.isalpha() and len(tl) <= 4 and tl not in known:
                    ev_d = _highlight_abbrev(t_desc, tl) if where_label == "desc3" else ""
                    ev_w = _highlight_abbrev(t_www, tl)  if where_label == "www"   else ""
                    abbr_occ.append((
                        rowp, where_label, tl,
                        ev_d, ev_w,
                        r.get(prod_col,""), r.get(sdn_col,""), r.get(attr_col,""),
                        r.get(d1_col,""), r.get(d2_col,""), r.get(d3src_col,""),
                        r.get(look_col,"")
                    ))
        collect_abbrevs(toks_desc, "desc3")
        collect_abbrevs(toks_www, "www")

        uniq = set([t.lower() for t in set(toks_desc + toks_www)
                    if t.isalpha() and len(t) <= 4 and t.lower() not in known])
        for t in uniq:
            abbr_rowu.append((rowp, t))

    # ---- Aggregate GLUE ----
    if glue_rows:
        glue_df = pd.DataFrame(glue_rows, columns=[
            "rowpointer","where","pair","concat",
            "evidence_desc3","evidence_www",
            "sample_prod","sample_source-desc-name","sample_attributegrp",
            "sample_descrip1","sample_descrip2","sample_descrip3","sample_lookup"
        ])
        grp = glue_df.groupby(["pair","concat"], as_index=False)
        gsum = grp.agg(
            occurrences=("rowpointer","count"),
            row_hits=("rowpointer", pd.Series.nunique),
            where=("where", lambda s: "both" if set(s)=={"desc3","www"} else (list(set(s))[0] if len(set(s))==1 else ",".join(sorted(set(s))))),
            evidence_desc3=("evidence_desc3","first"),
            evidence_www=("evidence_www","first"),
            sample_prod=("sample_prod","first"),
            **{"sample_source-desc-name":("sample_source-desc-name","first")},
            sample_attributegrp=("sample_attributegrp","first"),
            sample_descrip1=("sample_descrip1","first"),
            sample_descrip2=("sample_descrip2","first"),
            sample_descrip3=("sample_descrip3","first"),
            sample_lookup=("sample_lookup","first"),
        )
        gsum["issue_type"] = "GLUE_CANDIDATE"
        gsum["count"] = gsum["occurrences"]
        gsum["option1"] = gsum.apply(lambda r: f"Add glue: '{r['pair']}' → '{r['concat']}'", axis=1)
        gsum["option2"] = "Keep split tokens"
        gsum["option3"] = "Add synonym for glued form"
        gsum["option4"] = "Defer / Ignore"
    else:
        gsum = pd.DataFrame(columns=BASE_COLS)

    # ---- Aggregate ABBREV ----
    if abbr_occ or abbr_rowu:
        occ = pd.DataFrame(abbr_occ, columns=[
            "rowpointer","where","abbrev",
            "evidence_desc3","evidence_www",
            "sample_prod","sample_source-desc-name","sample_attributegrp",
            "sample_descrip1","sample_descrip2","sample_descrip3","sample_lookup"
        ]) if abbr_occ else pd.DataFrame(columns=[
            "rowpointer","where","abbrev","evidence_desc3","evidence_www",
            "sample_prod","sample_source-desc-name","sample_attributegrp",
            "sample_descrip1","sample_descrip2","sample_descrip3","sample_lookup"
        ])
        rowu = pd.DataFrame(abbr_rowu, columns=["rowpointer","abbrev"]) if abbr_rowu else pd.DataFrame(columns=["rowpointer","abbrev"])

        occ_grp = occ.groupby(["abbrev"], as_index=False).agg(
            occurrences=("rowpointer","count"),
            where=("where", lambda s: "both" if set(s)=={"desc3","www"} else (list(set(s))[0] if len(set(s))==1 else ",".join(sorted(set(s))))),
            evidence_desc3=("evidence_desc3","first"),
            evidence_www=("evidence_www","first"),
            sample_prod=("sample_prod","first"),
            **{"sample_source-desc-name":("sample_source-desc-name","first")},
            sample_attributegrp=("sample_attributegrp","first"),
            sample_descrip1=("sample_descrip1","first"),
            sample_descrip2=("sample_descrip2","first"),
            sample_descrip3=("sample_descrip3","first"),
            sample_lookup=("sample_lookup","first"),
        ) if not occ.empty else pd.DataFrame(columns=["abbrev","occurrences","where","evidence_desc3","evidence_www",
                                                      "sample_prod","sample_source-desc-name","sample_attributegrp",
                                                      "sample_descrip1","sample_descrip2","sample_descrip3","sample_lookup"])
        hits_grp = rowu.groupby(["abbrev"], as_index=False).agg(row_hits=("rowpointer", pd.Series.nunique)) if not rowu.empty else pd.DataFrame(columns=["abbrev","row_hits"])

        asum = pd.merge(occ_grp, hits_grp, on="abbrev", how="outer").fillna("")
        if "occurrences" not in asum.columns: asum["occurrences"] = 0
        if "row_hits" not in asum.columns:    asum["row_hits"] = 0
        if "where" not in asum.columns:       asum["where"] = ""
        asum["issue_type"] = "UNKNOWN_ABBREV"
        asum["count"] = asum["row_hits"]  # historical 'count' = row_hits for abbrevs
        asum["option1"] = asum.apply(lambda r: f"Add synonym: '{r['abbrev']}' → '<full term>'", axis=1)
        asum["option2"] = "Keep as-is"
        asum["option3"] = "Remove from desc3/www"
        asum["option4"] = "Defer / Ignore"
    else:
        asum = pd.DataFrame(columns=BASE_COLS)

    recom = pd.concat([gsum, asum], ignore_index=True, sort=False)
    if recom.empty:
        recom = pd.DataFrame(columns=BASE_COLS)

    # add decision columns
    recom["decision"] = ""
    recom["approved_value"] = ""
    recom["scope_hint"] = "line"
    recom["notes"] = ""

    ts = _now_ts()
    base = f"QA_Recom_Actn_en_{args.line}_{ts}"
    out_csv  = os.path.join(DECISIONS_DIR, f"{base}.csv")
    out_xlsx = os.path.join(DECISIONS_DIR, f"{base}.xlsx")

    recom = _ensure_decision_cols(recom)
    _write_csv(recom, out_csv)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xlw:
        recom.to_excel(xlw, index=False)
    _xlsx_format_template(out_xlsx)

    print("Decision workbook written:")
    print(" ", out_csv)
    print(" ", out_xlsx)

# ---------- SUMMARY / VALIDATE / APPLY ----------
def _conflict_scan(df: pd.DataFrame, line: str) -> pd.DataFrame:
    line_cfg = _load_json(LINE_FILE_TMPL.format(line=line)) or {}
    line_syn = {**(line_cfg.get("synonyms") or {}), **(line_cfg.get("pricing_synonyms") or {})}
    line_glu = (line_cfg.get("glue_bigrams") or {})
    glob_syn = _load_json(GLOBAL_SYNONYMS) or {}
    glob_glu = _load_json(GLOBAL_GLUE) or {}

    issues = []
    for _, r in df.iterrows():
        decision = (r.get("decision","") or "").strip().lower()
        if decision not in ("accept_line","accept_global"):
            continue
        it = (r.get("issue_type","") or "").strip().upper()
        if it == "GLUE_CANDIDATE":
            pair = (r.get("pair","") or "").strip().lower()
            target = (r.get("approved_value","") or r.get("concat","") or "").strip().lower()
            if not pair or not target: continue
            if line_glu.get(pair) and line_glu.get(pair) != target:
                issues.append(("GLUE", "line_conflict", pair, line_glu.get(pair), target))
            if glob_glu.get(pair) and glob_glu.get(pair) != target:
                issues.append(("GLUE", "global_conflict", pair, glob_glu.get(pair), target))
        elif it == "UNKNOWN_ABBREV":
            abbr = (r.get("abbrev","") or "").strip().lower()
            full = (r.get("approved_value","") or "").strip().lower()
            if not abbr or not full: continue
            if line_syn.get(abbr) and line_syn.get(abbr) != full:
                issues.append(("SYN", "line_conflict", abbr, line_syn.get(abbr), full))
            if glob_syn.get(abbr) and glob_syn.get(abbr) != full:
                issues.append(("SYN", "global_conflict", abbr, glob_syn.get(abbr), full))
    return pd.DataFrame(issues, columns=["type","scope","key","existing","proposed"]) if issues else pd.DataFrame(columns=["type","scope","key","existing","proposed"])

def cmd_summary(args):
    in_path = args.input
    if not in_path:
        # latest decisions for line
        patt = os.path.join(DECISIONS_DIR, f"QA_Recom_Actn_en_{args.line}_*.xlsx")
        cand = sorted(glob.glob(patt))
        if cand:
            in_path = cand[-1]
        else:
            patt = os.path.join(DECISIONS_DIR, f"QA_Recom_Actn_en_{args.line}_*.csv")
            cand = sorted(glob.glob(patt))
            if cand: in_path = cand[-1]
    if not in_path:
        sys.exit("summary: no input decisions file found")

    df = _read_any(in_path)
    total = len(df)
    by_type = df["issue_type"].value_counts(dropna=False)
    by_dec  = df["decision"].fillna("").replace("", "<blank>").value_counts(dropna=False)

    print(f"Summary for {in_path}")
    print(f"Total rows: {total}")
    print("\nBy issue_type:\n", by_type.to_string())
    print("\nBy decision:\n", by_dec.to_string())

    confl = _conflict_scan(df, args.line)
    summo = os.path.join(QA_DIR_DEFAULT, f"QA_Recom_Summary_en_{args.line}_{_now_ts()}.txt")
    with open(summo, "w", encoding="utf-8") as f:
        f.write(f"Input: {in_path}\nTotal rows: {total}\n\nBy issue_type:\n{by_type.to_string()}\n\nBy decision:\n{by_dec.to_string()}\n")
        if not confl.empty:
            f.write("\nConflicts detected:\n")
            f.write(confl.to_string(index=False))
    print(f"\nSummary written: {summo}")

    if not confl.empty:
        confcsv = os.path.join(QA_DIR_DEFAULT, f"QA_Recom_Conflicts_en_{args.line}_{_now_ts()}.csv")
        _write_csv(confl, confcsv)
        print(f"Conflicts CSV: {confcsv}")
    else:
        print("No conflicts detected ✅")

def cmd_validate(args):
    if not args.input:
        sys.exit("--in (path to decisions CSV/XLSX) is required for validate")
    df = _read_any(args.input)
    errs = []
    for i, r in df.iterrows():
        dec = (r.get("decision","") or "").strip().lower()
        if dec and dec not in DECISION_ALLOWED:
            errs.append(f"Row {i+2}: decision '{dec}' not in {DECISION_ALLOWED}")
        if dec.startswith("accept"):
            it = (r.get("issue_type","") or "").strip().upper()
            if it == "UNKNOWN_ABBREV":
                av = (r.get("approved_value","") or "").strip()
                if not av:
                    errs.append(f"Row {i+2}: approved_value required for UNKNOWN_ABBREV when decision is '{dec}'")
    if errs:
        print("Validation FAILED ❌")
        for e in errs: print(" -", e)
        sys.exit(2)
    print("Validation PASSED ✅")

def _snapshot(label: str, data: dict) -> str:
    _ensure_dir(SNAP_DIR)
    ts = _now_ts()
    out = os.path.join(SNAP_DIR, f"{label}_{ts}.json")
    _save_json(out, data); return out

def cmd_apply(args):
    if not args.input:
        sys.exit("--in (path to decisions CSV/XLSX) is required for apply")
    df = _read_any(args.input)

    line_path = LINE_FILE_TMPL.format(line=args.line)
    line_cfg  = _load_json(line_path) or {}
    line_cfg.setdefault("synonyms", {})
    line_cfg.setdefault("glue_bigrams", {})

    glob_syn = _load_json(GLOBAL_SYNONYMS) or {}
    glob_glu = _load_json(GLOBAL_GLUE) or {}

    applied = {"line_syn":0, "line_glue":0, "glob_syn":0, "glob_glue":0}
    notes   = []

    for i, r in df.iterrows():
        decision = (r.get("decision","") or "").strip().lower()
        if decision not in ("accept_line","accept_global"):
            continue
        it = (r.get("issue_type","") or "").strip().upper()

        if it == "GLUE_CANDIDATE":
            pair = (r.get("pair","") or "").strip().lower()
            target = (r.get("approved_value","") or r.get("concat","") or "").strip().lower()
            if not pair or not target:
                notes.append(f"Row {i+2}: skipped GLUE (missing pair/target)")
                continue
            if decision == "accept_line":
                line_cfg["glue_bigrams"][pair] = target
                applied["line_glue"] += 1
            else:
                glob_glu[pair] = target
                applied["glob_glue"] += 1

        elif it == "UNKNOWN_ABBREV":
            abbr = (r.get("abbrev","") or "").strip().lower()
            full = (r.get("approved_value","") or "").strip().lower()
            if not abbr or not full:
                notes.append(f"Row {i+2}: skipped SYN (missing abbrev/approved_value)")
                continue
            if decision == "accept_line":
                line_cfg["synonyms"][abbr] = full
                applied["line_syn"] += 1
            else:
                glob_syn[abbr] = full
                applied["glob_syn"] += 1

    snap_line_syn = _snapshot(f"line_synonyms_{args.line}", line_cfg.get("synonyms", {}))
    snap_line_glu = _snapshot(f"line_glue_{args.line}",      line_cfg.get("glue_bigrams", {}))
    snap_glob_syn = _snapshot("global_synonyms",             glob_syn)
    snap_glob_glu = _snapshot("global_glue",                 glob_glu)

    _save_json(line_path, line_cfg)
    _save_json(GLOBAL_SYNONYMS, glob_syn)
    _save_json(GLOBAL_GLUE,     glob_glu)

    report = os.path.join(QA_DIR_DEFAULT, f"QA_Recom_ApplyReport_en_{args.line}_{_now_ts()}.txt")
    with open(report, "w", encoding="utf-8") as f:
        f.write("Applied counts:\n")
        for k,v in applied.items(): f.write(f"  {k}: {v}\n")
        f.write("\nNotes:\n")
        for n in notes: f.write(f" - {n}\n")
        f.write("\nSnapshots:\n")
        for p in [snap_line_syn, snap_line_glu, snap_glob_syn, snap_glob_glu]:
            f.write(f"  {p}\n")
        f.write("\nUpdated configs:\n")
        f.write(f"  {line_path}\n  {GLOBAL_SYNONYMS}\n  {GLOBAL_GLUE}\n")
    print(f"Apply complete. Report: {report}")
    print("Counts:", applied)

# ---------- CLI ----------
def main():
    ap = argparse.ArgumentParser(description="Decision kit (EN) for desc_src_www and descrip3 highlighting + rule promotion.")
    sp = ap.add_subparsers(dest="cmd", required=True)

    # DERIVE
    ap_d = sp.add_parser("derive", help="Create decisions from enhanced QA (QA_<line>_en_*.csv) with evidence highlighting.")
    ap_d.add_argument("--line", required=True)
    ap_d.add_argument("--in", dest="input", help="Path to an enhanced QA CSV/XLSX (default: newest in rep/qa).")
    ap_d.set_defaults(func=cmd_derive)

    # SUMMARY
    ap_s = sp.add_parser("summary", help="Summarize decisions and list conflicts against existing configs.")
    ap_s.add_argument("--line", required=True)
    ap_s.add_argument("--in", dest="input", help="Path to decisions CSV/XLSX (default: latest decisions for line).")
    ap_s.set_defaults(func=cmd_summary)

    # VALIDATE
    ap_v = sp.add_parser("validate", help="Validate decisions before apply.")
    ap_v.add_argument("--line", required=True)
    ap_v.add_argument("--in", dest="input", required=True)
    ap_v.set_defaults(func=cmd_validate)

    # APPLY
    ap_a = sp.add_parser("apply", help="Apply accepted rules to line/global configs, with snapshots.")
    ap_a.add_argument("--line", required=True)
    ap_a.add_argument("--in", dest="input", required=True)
    ap_a.set_defaults(func=cmd_apply)

    args = ap.parse_args()
    args.func(args)

if __name__ == "__main__":
    main()
